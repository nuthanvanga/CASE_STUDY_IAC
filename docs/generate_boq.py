"""
Generates a multi-sheet Excel BOQ from the line items defined below.
Run from /sessions/intelligent-kind-lamport/mnt/outputs/azure-prod-iac/docs/.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WB_PATH = "BOQ_UAE_North.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFD966")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_sheet(ws, title, headers, rows, total_label="Subtotal"):
    ws.title = title
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for r in rows:
        ws.append(r)
    last_row = ws.max_row
    # Subtotal row
    subtotal_formula = f"=SUM({get_column_letter(len(headers))}2:{get_column_letter(len(headers))}{last_row})"
    ws.append([total_label] + [""] * (len(headers) - 2) + [subtotal_formula])
    sr = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=sr, column=col_idx)
        c.fill = SUBTOTAL_FILL
        c.font = Font(bold=True)
        c.border = BORDER
    # Number format on monthly column
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=len(headers))
        cell.number_format = '"$"#,##0.00'
    # Borders + widths
    for r in range(1, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=col_idx).border = BORDER
    widths = [6, 42, 36, 10, 18, 16]
    for i, w in enumerate(widths[:len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return sr  # subtotal row index

def main():
    wb = Workbook()

    # ----------------------------------------------------------- Compute
    compute = wb.active
    headers = ["#", "Resource", "SKU / Tier", "Qty", "Unit price (USD)", "Monthly (USD)"]
    rows = [
        [1, "AKS control plane", "Standard tier (Uptime SLA)", 1, "$0.10/hr", 73.00],
        [2, "AKS system node pool", "Standard_D4s_v5 (4 vCPU/16 GB)", 3, "$0.232/hr", 508.00],
        [3, "AKS user node pool", "Standard_D8s_v5 (8 vCPU/32 GB)", 3, "$0.464/hr", 1016.00],
        [4, "App Service Plan", "Premium V3 P1v3, zone-redundant", 3, "$0.214/hr", 469.00],
        [5, "App Service staging slot", "shared on the plan", 1, "included", 0.00],
        [6, "App Insights ingestion", "ingestion", 5, "$2.30/GB", 12.00],
    ]
    write_sheet(compute, "Compute", headers, rows)

    # ----------------------------------------------------------- Containers
    containers = wb.create_sheet()
    rows = [
        [1, "Azure Container Registry", "Premium", 1, "$1.667/day", 50.00],
        [2, "ACR storage above 500 GB", "per GB", 100, "$0.10/GB", 10.00],
        [3, "ACR geo-replication (UAE Central)", "secondary region", 1, "$1.667/day", 50.00],
    ]
    write_sheet(containers, "Containers", headers, rows)

    # ----------------------------------------------------------- Networking
    networking = wb.create_sheet()
    rows = [
        [1, "Virtual Network", "included", 1, "$0", 0.00],
        [2, "NAT Gateway (zone-redundant)", "per hour", 1, "$0.045/hr", 33.00],
        [3, "NAT Gateway data processed", "per GB", 1000, "$0.045/GB", 45.00],
        [4, "Public IP Standard (static)", "per hour", 2, "$0.005/hr", 7.30],
        [5, "Private endpoints (KV, ACR)", "per endpoint/hr", 2, "$0.01/hr", 14.60],
        [6, "Private endpoint data processed", "per GB", 500, "$0.01/GB", 5.00],
        [7, "Private DNS zones", "per zone", 2, "$0.50/zone/mo", 1.00],
        [8, "Application Gateway v2 (WAF)", "small WAFv2", 1, "~$200/mo", 200.00],
        [9, "Egress bandwidth (after 100 GB free)", "per GB", 500, "$0.087/GB", 43.00],
        [10, "Cross-region replication egress", "per GB", 400, "$0.087/GB", 35.00],
    ]
    write_sheet(networking, "Networking", headers, rows)

    # ----------------------------------------------------------- Security
    security = wb.create_sheet()
    rows = [
        [1, "Key Vault (Premium) transactions", "per 10K ops", 300, "$0.03/10K", 9.00],
        [2, "Key Vault HSM keys", "per key/month", 5, "$5.00", 25.00],
        [3, "Defender for Cloud (Foundational CSPM)", "free tier", 1, "$0", 0.00],
        [4, "Defender for Containers", "per vCore/hr", 30, "$0.0095/hr", 208.00],
        [5, "Defender for App Service", "per node/hr", 3, "$0.02/hr", 43.80],
        [6, "Microsoft Sentinel (optional)", "per GB ingested", 30, "$2.46/GB", 74.00],
    ]
    write_sheet(security, "Security", headers, rows)

    # ----------------------------------------------------------- Monitoring
    monitoring = wb.create_sheet()
    rows = [
        [1, "Log Analytics ingestion", "PAYG", 100, "$2.76/GB", 276.00],
        [2, "Log Analytics extended retention", "per GB/mo (2 mo)", 200, "$0.10/GB", 20.00],
        [3, "Container Insights", "included with LAW", 1, "$0", 0.00],
        [4, "App Insights ingestion (App Svc)", "ingestion", 30, "$2.30/GB", 69.00],
        [5, "Azure Monitor metric alerts", "per signal", 50, "$0.10/signal", 5.00],
        [6, "Managed Prometheus (extra samples)", "per 10M samples", 90, "$0.16/10M", 14.40],
        [7, "Azure Managed Grafana", "Standard", 1, "$0.105/hr", 76.65],
        [8, "Diagnostic settings", "included", 1, "$0", 0.00],
    ]
    write_sheet(monitoring, "Monitoring", headers, rows)

    # ----------------------------------------------------------- Storage
    storage = wb.create_sheet()
    rows = [
        [1, "Premium ZRS managed disks (PVCs)", "P15 (256 GiB)", 3, "$35/disk", 105.00],
        [2, "Snapshot storage (PVC backup)", "LRS standard", 100, "$0.05/GB", 5.00],
        [3, "App Service backup", "included w/ Premium", 1, "$0", 0.00],
    ]
    write_sheet(storage, "Storage_Backup", headers, rows)

    # ----------------------------------------------------------- Summary
    summary = wb.create_sheet("Summary", 0)
    summary.append(["Bill of Quantity — Production Environment, UAE North"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Region", "UAE North"])
    summary.append(["Currency", "USD"])
    summary.append(["Pricing reference", "Azure public retail (May 2026 snapshot)"])
    summary.append([])
    summary.append(["Category", "Monthly (PAYG)"])
    for c in summary[summary.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    cats = [
        ("Compute", "=Compute!F" + str(compute.max_row)),
        ("Containers & Registry", "=Containers!F" + str(containers.max_row)),
        ("Networking", "=Networking!F" + str(networking.max_row)),
        ("Security", "=Security!F" + str(security.max_row)),
        ("Monitoring & Logging", "=Monitoring!F" + str(monitoring.max_row)),
        ("Storage & Backup", "=Storage_Backup!F" + str(storage.max_row)),
    ]
    for cat, formula in cats:
        summary.append([cat, formula])
    total_row = summary.max_row + 1
    summary.append(["Total (PAYG estimate)", f"=SUM(B8:B{total_row - 1})"])
    for c in summary[total_row]:
        c.fill = TOTAL_FILL
        c.font = Font(bold=True)
    for r in range(8, total_row + 1):
        summary.cell(row=r, column=2).number_format = '"$"#,##0.00'
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 22
    summary.append([])
    summary.append(["3-year Reservation savings (compute only)", "approx. 45–55% off compute & app service"])
    summary.append(["Recommended budget (with 10% buffer)", f"=ROUND(B{total_row}*1.1, 2)"])
    summary.cell(row=summary.max_row, column=2).number_format = '"$"#,##0.00'

    wb.save(WB_PATH)
    print(f"Wrote {WB_PATH}")

if __name__ == "__main__":
    main()
